import sys
import os
import re
import json
import time
import random
import shutil
import traceback
from datetime import datetime, timedelta, timezone
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QDateEdit, QTableWidget,
                             QTableWidgetItem, QHeaderView, QMessageBox, QFileDialog,
                             QTextEdit, QGroupBox, QSplitter, QDialog, QFormLayout, QComboBox, QCheckBox)
from PyQt5.QtCore import QDate, Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from bilibili_api import user, sync, Credential, request_settings


# -------------------------- 工具函数：获取用户数据目录 --------------------------
def get_app_data_dir():
    home_dir = os.path.expanduser("~")
    app_dir = os.path.join(home_dir, ".BiliStatTool")
    if not os.path.exists(app_dir):
        os.makedirs(app_dir)
    return app_dir


# ====================== 【零缺陷】7天统计数据 本地存储+自动备份 ======================
def get_daily_data_path():
    return os.path.join(get_app_data_dir(), "video_7day_data.json")


def get_backup_dir():
    return os.path.join(get_app_data_dir(), "backups")


# 自动清理旧备份（保留30天，可修改）
def clean_old_backups(retention_days=30):
    backup_dir = get_backup_dir()
    if not os.path.exists(backup_dir):
        return
    now = time.time()
    max_age = retention_days * 86400
    for f in os.listdir(backup_dir):
        path = os.path.join(backup_dir, f)
        try:
            if os.path.isfile(path) and (now - os.path.getmtime(path)) > max_age:
                os.remove(path)
        except:
            continue


# 原子化保存 + 自动备份（永不覆盖、永不损坏）
def save_daily_video_data(data):
    data_path = get_daily_data_path()
    backup_dir = get_backup_dir()
    os.makedirs(backup_dir, exist_ok=True)

    # 1. 原子写入主数据（防止中途崩溃损坏）
    temp_path = data_path + ".tmp"
    try:
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        # 替换原文件（原子操作）
        if os.path.exists(data_path):
            os.replace(temp_path, data_path)
        else:
            os.rename(temp_path, data_path)
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise e

    # 2. 创建带时间戳的唯一备份（绝不覆盖）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"video_7day_backup_{timestamp}.json")
    try:
        shutil.copy2(data_path, backup_path)
    except:
        pass

    # 3. 自动清理30天前的旧备份
    clean_old_backups(30)


def load_daily_video_data():
    path = get_daily_data_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                # 自动迁移：为旧数据增加 excluded 字段
                for bvid, v in data.items():
                    if "excluded" not in v:
                        v["excluded"] = False
                return data
        except:
            return {}
    return {}


# ============================================================================


def load_cookie_config():
    cookie_file = os.path.join(get_app_data_dir(), "cookie.json")
    if os.path.exists(cookie_file):
        try:
            with open(cookie_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    # 默认值（新增两个字段）
    return {
        "SESSDATA": "",
        "BILI_JCT": "",
        "BUVID3": "",
        "DEDEUSERID": "",
        "AC_TIME_VALUE": "1000"
    }


def save_cookie_config(sessdata, bili_jct, buvid3, dedeuserid, ac_time_value):
    cookie_file = os.path.join(get_app_data_dir(), "cookie.json")
    with open(cookie_file, "w", encoding="utf-8") as f:
        json.dump({
            "SESSDATA": sessdata,
            "BILI_JCT": bili_jct,
            "BUVID3": buvid3,
            "DEDEUSERID": dedeuserid,
            "AC_TIME_VALUE": ac_time_value
        }, f, ensure_ascii=False, indent=2)


# -------------------------- 全局配置 --------------------------
cookie_config = load_cookie_config()
SESSDATA = cookie_config.get("SESSDATA", "")
BILI_JCT = cookie_config.get("BILI_JCT", "")
BUVID3 = cookie_config.get("BUVID3", "")
DEDEUSERID = cookie_config.get("DEDEUSERID", "")  # 新增
AC_TIME_VALUE = cookie_config.get("AC_TIME_VALUE", "1000")  # 新增，默认 1000 就行

MAX_UP_COUNT = 80
CONFIG_FILE = os.path.join(get_app_data_dir(), "up_list.json")
TIMEZONE_CN = timezone(timedelta(hours=8), "Asia/Shanghai")

RANK_LEVEL_CONFIG = [
    ("TOP1", 1, 1, "D6C7FB"),
    ("TOP2-3", 2, 3, "FEDBF5"),
    ("TOP4-5", 4, 5, "FEF7FF"),
    ("TOP6-10", 6, 10, "E0EEFF"),
    ("TOP11-15", 11, 15, "FFFDE5"),
    ("TOP16-20", 16, 20, "F5F891"),
    ("TOP21-25", 21, 25, "D1F9BA"),
    ("TOP26-30", 26, 30, "72A1A7"),
    ("TOP31-35", 31, 35, "BA9F7D"),
    ("TOP36-40", 36, 40, "C4D69F"),
    ("TOP41-50", 41, 50, "FFBE4E"),
    ("TOP50+", 51, 999, "9e9e9e"),
]
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
content_font = Font(name='微软雅黑', size=10)
header_fill = PatternFill('solid', fgColor='4472C4')
odd_row_fill = PatternFill('solid', fgColor='F2F2F2')
even_row_fill = PatternFill('solid', fgColor='FFFFFF')


# -------------------------- 统计线程（已完全重构 + 真人行为模拟 + 防412） --------------------------
class StatThread(QThread):
    log_signal = pyqtSignal(str)
    table_signal = pyqtSignal(list)
    finish_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)

    def __init__(self, up_list, start_date, end_date,
                 sessdata, bili_jct, buvid3, dedeuserid, ac_time_value,
                 proxy=None):
        super().__init__()
        self.up_list = up_list
        self.start_date = start_date
        self.end_date = end_date
        self.sessdata = sessdata
        self.bili_jct = bili_jct
        self.buvid3 = buvid3
        self.dedeuserid = dedeuserid  # 新增
        self.ac_time_value = ac_time_value  # 新增
        self.proxy = proxy
        self.current_page = 0
        self.daily_data = load_daily_video_data()

    def _human_like_delay(self, delay_type="page"):
        """
        模拟真人操作的不规则延时
        delay_type:
            - up: 切换UP主的长停顿（3-8秒，10%概率触发15-30秒长休息）
            - page: 翻页的中等停顿（2-6秒）
            - retry: 重试前的超长冷却（30-60秒）
        """
        if delay_type == "up":
            if random.random() < 0.1:
                delay = random.uniform(15, 30)
                self.log_signal.emit(f"  ☕ 模拟真人休息 {delay:.1f} 秒...")
            else:
                delay = random.uniform(3, 12)
                self.log_signal.emit(f"  ↳ 切换UP主，等待 {delay:.1f} 秒...")
        elif delay_type == "page":
            delay = random.uniform(40, 120)
            self.log_signal.emit(f"  ↳ 第 {self.current_page} 页完成，等待 {delay:.1f} 秒...")
        elif delay_type == "retry":
            delay = random.uniform(60, 120)
            self.log_signal.emit(f"  🚨 触发风控，强制冷却 {delay:.1f} 秒后重试...")
        time.sleep(delay)

    def _process_video(self, video, uid, nickname, now):
        """统一处理单个视频逻辑（消除重复代码 + 保证 daily_data 始终更新）"""
        pub_time = datetime.fromtimestamp(video["created"], tz=TIMEZONE_CN)
        if pub_time < self.start_date or pub_time > self.end_date:
            return None

        bvid = video["bvid"]
        title = re.sub(r'[\r\n\t]+', ' ', video["title"]).strip()
        days_diff = (now - pub_time).days

        if days_diff >= 7:
            if bvid in self.daily_data:
                current_play = self.daily_data[bvid]["final_play"]
                play_tag = "已结算"
                stat_days = 7
            else:
                play_raw = video.get("play", 0)
                current_play = int(play_raw) if str(play_raw).isdigit() else 0
                play_tag = "已结算"
                stat_days = 7
                self.daily_data[bvid] = {
                    "uid": uid, "nickname": nickname, "title": title,
                    "pub_time": pub_time.strftime("%Y-%m-%d %H:%M"),
                    "bvid": bvid, "final_play": current_play,
                    "stat_days": stat_days, "status": play_tag
                }
        else:
            play_raw = video.get("play", 0)
            current_play = int(play_raw) if str(play_raw).isdigit() else 0
            play_tag = "统计中"
            stat_days = days_diff + 1
            self.daily_data[bvid] = {
                "uid": uid, "nickname": nickname, "title": title,
                "pub_time": pub_time.strftime("%Y-%m-%d %H:%M"),
                "bvid": bvid, "final_play": current_play,
                "stat_days": stat_days, "status": play_tag
            }

        return {
            "uid": uid, "nickname": nickname, "title": title, "bvid": bvid,
            "pub_time": pub_time.strftime("%Y-%m-%d %H:%M"),
            "current_play": current_play, "stat_days": stat_days,
            "play_tag": play_tag
        }

    def run(self):
        # ==================== 防412风控设置 ====================
        try:
            if self.proxy:
                request_settings.set_proxy(self.proxy)
                self.log_signal.emit(f"✅ 已启用代理：{self.proxy}")
            else:
                request_settings.set_proxy(None)

            request_settings.set("impersonate", "chrome131")
            request_settings.set("http2", True)
            self.log_signal.emit("✅ 已启用浏览器伪装（chrome131）")
        except Exception as e:
            self.log_signal.emit(f"⚠️ 代理/伪装设置失败：{e}")

        cred = None
        if self.sessdata and self.bili_jct and self.buvid3:
            try:
                cred = Credential(
                    sessdata=self.sessdata,
                    bili_jct=self.bili_jct,
                    buvid3=self.buvid3,
                    dedeuserid=self.dedeuserid,
                    ac_time_value=self.ac_time_value
                )
            except Exception as e:
                self.log_signal.emit(f"⚠️ Credential 创建失败：{e}")
        else:
            self.log_signal.emit("⚠️ Cookie 不完整，将使用游客模式")

        result_data = []
        up_total_list = []
        now = datetime.now(tz=TIMEZONE_CN)

        # 【关键优化1】随机打乱UP主请求顺序，避免固定规律
        # random.shuffle(self.up_list)
        # for up in self.up_list:

        # 改成
        up_list_copy = self.up_list[:]  # 或者 list(self.up_list)
        random.shuffle(up_list_copy)
        for up in up_list_copy:  # 后面所有地方改用 up_list_copy
            if self.isInterruptionRequested():
                self.log_signal.emit("统计已被用户中断")
                save_daily_video_data(self.daily_data)
                return

            uid = int(up["uid"])
            nickname = up["nickname"]
            self.log_signal.emit(f"正在统计：{nickname}（UID:{uid}）")
            self.current_page = 0

            # 【关键优化2】切换UP主前，先加一次长停顿
            self._human_like_delay(delay_type="up")

            up_video_list = []
            up_total_play = 0
            page = 1
            has_more = True

            try:
                u = user.User(uid, credential=cred) if cred else user.User(uid)

                # 【关键优化3】动态设置Referer为当前UP主的个人主页
                request_settings.set("headers", {
                    "Referer": f"https://space.bilibili.com/{uid}/video",
                    "Origin": "https://space.bilibili.com",
                    "Accept": "application/json, text/plain, */*",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Cache-Control": "no-cache",
                    "Pragma": "no-cache"
                })

                while has_more:
                    if self.isInterruptionRequested():
                        save_daily_video_data(self.daily_data)
                        return

                    res = sync(u.get_videos(pn=page, ps=30))
                    video_list = res.get("list", {}).get("vlist", [])
                    if not video_list:
                        has_more = False
                        break

                    # 【关键优化4】提前判断：当前页最早的视频已经早于开始日期，直接终止翻页
                    earliest_time = datetime.fromtimestamp(video_list[-1]["created"], tz=TIMEZONE_CN)
                    if earliest_time < self.start_date:
                        has_more = False
                        # 只处理当前页里晚于开始日期的视频，不用继续翻页
                        for video in video_list:
                            processed = self._process_video(video, uid, nickname, now)
                            if processed:
                                up_video_list.append(processed)
                                up_total_play += processed["current_play"]
                        break

                    # 正常处理当前页
                    for video in video_list:
                        processed = self._process_video(video, uid, nickname, now)
                        if processed:
                            up_video_list.append(processed)
                            up_total_play += processed["current_play"]

                    if len(video_list) < 30:
                        has_more = False
                    page += 1
                    self.current_page = page

                    # 【关键优化5】使用真人级不规则延时
                    self._human_like_delay(delay_type="page")

            except Exception as e:
                err_str = str(e)
                if "412" in err_str or "Precondition Failed" in err_str:
                    self.log_signal.emit(f"❌ {nickname} 触发风控（412），先强制冷却...")
                    # 【关键优化6】触发412后，先超长冷却，而不是立即重试
                    self._human_like_delay(delay_type="retry")
                    self.log_signal.emit(f"🔄 尝试游客模式获取 {nickname} 的投稿...")
                else:
                    self.error_signal.emit(f"获取 {nickname} 投稿失败：{err_str}")
                    continue

                # 游客模式重试逻辑（仅冷却后执行）
                try:
                    u = user.User(uid)
                    page = 1
                    has_more = True
                    temp_videos = []
                    while has_more:
                        if self.isInterruptionRequested():
                            self.log_signal.emit("统计已被用户中断")
                            save_daily_video_data(self.daily_data)
                            return

                        res = sync(u.get_videos(pn=page, ps=30))
                        video_list = res.get("list", {}).get("vlist", [])
                        if not video_list:
                            break

                        earliest_time = datetime.fromtimestamp(video_list[-1]["created"], tz=TIMEZONE_CN)
                        if earliest_time < self.start_date:
                            has_more = False
                            # 处理当前页符合条件的视频
                            for video in video_list:
                                processed = self._process_video(video, uid, nickname, now)
                                if processed:
                                    temp_videos.append(processed)
                            break

                        # 正常处理当前页
                        for video in video_list:
                            processed = self._process_video(video, uid, nickname, now)
                            if processed:
                                temp_videos.append(processed)

                        if len(video_list) < 30:
                            has_more = False
                        page += 1
                        self.current_page = page
                        self._human_like_delay(delay_type="page")

                    up_video_list = temp_videos
                    up_total_play = sum(v["current_play"] for v in temp_videos)
                    self.log_signal.emit(f"✅ {nickname} 游客模式获取成功")
                except Exception as e2:
                    self.error_signal.emit(f"游客模式也失败：{str(e2)}")
                    continue

            result_data.extend(up_video_list)
            up_total_list.append({
                "nickname": nickname, "uid": uid, "total_play": up_total_play
            })
            self.table_signal.emit(result_data)

        # 统计完成，保存所有数据
        save_daily_video_data(self.daily_data)
        up_total_list_sorted = sorted(up_total_list, key=lambda x: (-x["total_play"], x["nickname"]))
        for idx, item in enumerate(up_total_list_sorted, 1):
            item["rank"] = idx
        self.finish_signal.emit(up_total_list_sorted)


# -------------------------- 作品筛选对话框（带UP主专属视图） --------------------------
# -------------------------- 作品筛选对话框（完整修复版） --------------------------
class VideoSelectionDialog(QDialog):
    def __init__(self, current_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("筛选统计作品")
        self.setModal(True)
        self.resize(1200, 750)
        self.current_data = current_data  # 本次统计的数据
        self.data_list = []  # 存储所有数据 [{data:..., selected:...}, ...]

        # 先初始化UI
        self.init_ui()
        # 再加载数据
        self.load_and_refresh()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        tip = QLabel("提示：取消勾选的视频将从排行榜和明细表中永久排除（历史数据也会生效）")
        tip.setStyleSheet("color: #D32F2F; font-weight:bold;")
        layout.addWidget(tip)

        # 顶部控制栏
        ctrl_layout = QHBoxLayout()

        # 1. 历史数据开关
        self.history_check = QCheckBox("包含全部历史数据一起筛选（推荐）")
        self.history_check.setChecked(True)
        self.history_check.stateChanged.connect(self.load_and_refresh)
        ctrl_layout.addWidget(self.history_check)

        # 2. 显示已排除开关
        self.show_excluded_check = QCheckBox("显示已排除的视频（用于恢复）")
        self.show_excluded_check.setChecked(False)
        self.show_excluded_check.stateChanged.connect(self.load_and_refresh)
        ctrl_layout.addWidget(self.show_excluded_check)

        ctrl_layout.addStretch()
        layout.addLayout(ctrl_layout)

        # UP主筛选栏
        up_layout = QHBoxLayout()
        up_layout.addWidget(QLabel("UP主专属视图："))
        self.up_combo = QComboBox()
        self.up_combo.setMinimumWidth(320)
        self.up_combo.currentIndexChanged.connect(self.refresh_table_view)
        up_layout.addWidget(self.up_combo)
        up_layout.addStretch()
        layout.addLayout(up_layout)

        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["选择", "UP主昵称", "视频标题", "BV号", "发布时间", "7天播放量", "统计天数", "状态"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        # 底部按钮
        btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选当前显示")
        self.select_all_btn.clicked.connect(self.select_current_visible)
        btn_layout.addWidget(self.select_all_btn)

        self.deselect_all_btn = QPushButton("全不选当前显示")
        self.deselect_all_btn.clicked.connect(self.deselect_current_visible)
        btn_layout.addWidget(self.deselect_all_btn)
        btn_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        confirm_btn = QPushButton("确认筛选并永久排除")
        confirm_btn.setStyleSheet("background-color: #D32F2F; color:white;")
        confirm_btn.clicked.connect(self.confirm_selection)
        btn_layout.addWidget(confirm_btn)
        layout.addLayout(btn_layout)

    def load_and_refresh(self):
        """重新从磁盘加载数据并刷新一切"""
        # 1. 加载数据
        self.data_list = self._get_full_data_internal()

        # 2. 刷新UP主下拉框（阻塞信号防止递归）
        self.up_combo.blockSignals(True)
        self.up_combo.clear()
        self.up_combo.addItem("全部UP主", None)

        up_set = set()
        for item in self.data_list:
            d = item.get("data", {})
            uid = d.get("uid", 0)
            nick = d.get("nickname", "")
            if uid and nick:
                key = (uid, nick)
                if key not in up_set:
                    up_set.add(key)
                    self.up_combo.addItem(f"{nick}（UID:{uid}）", key)

        self.up_combo.blockSignals(False)

        # 3. 刷新表格视图
        self.refresh_table_view()

    def _get_full_data_internal(self):
        """内部函数：获取合并后的数据"""
        if not self.history_check.isChecked():
            # 仅本次数据模式
            result = []
            show_excluded = self.show_excluded_check.isChecked()
            for item in self.current_data:
                is_excluded = item.get("excluded", False)
                if not show_excluded and is_excluded:
                    continue
                result.append({
                    "data": item,
                    "selected": not is_excluded,
                    "source_bvid": item.get("bvid", "")
                })
            return result

        daily = load_daily_video_data()
        full = []
        seen_bvids = set()
        show_excluded = self.show_excluded_check.isChecked()

        # 1. 历史数据
        for v in daily.values():
            bvid = v.get("bvid", "")
            if not bvid or bvid in seen_bvids:
                continue

            is_excluded = v.get("excluded", False)
            if not show_excluded and is_excluded:
                continue

            seen_bvids.add(bvid)
            full.append({
                "data": {
                    "uid": v.get("uid", 0),
                    "nickname": v.get("nickname", ""),
                    "title": v.get("title", ""),
                    "bvid": bvid,
                    "pub_time": v.get("pub_time", ""),
                    "current_play": v.get("final_play", 0),
                    "stat_days": v.get("stat_days", 7),
                    "play_tag": v.get("status", "已结算"),
                },
                "selected": not is_excluded,
                "source_bvid": bvid
            })

        # 2. 本次新增数据
        for item in self.current_data:
            bvid = item.get("bvid", "")
            if not bvid or bvid in seen_bvids:
                continue

            is_excluded = item.get("excluded", False)
            if not show_excluded and is_excluded:
                continue

            seen_bvids.add(bvid)
            full.append({
                "data": item,
                "selected": not is_excluded,
                "source_bvid": bvid
            })
        return full

    def refresh_table_view(self):
        """仅刷新表格显示"""
        # 1. 断开信号防闪烁
        try:
            self.table.itemChanged.disconnect()
        except:
            pass

        # 2. 获取当前筛选的UP主
        current_up = self.up_combo.currentData()

        # 3. 填充表格
        self.table.setRowCount(0)
        row_pos = 0

        for idx, item in enumerate(self.data_list):
            d = item.get("data", {})

            # UP主过滤
            if current_up is not None:
                target_uid, _ = current_up
                if d.get("uid", 0) != target_uid:
                    continue

            # 插入行
            self.table.insertRow(row_pos)

            # 勾选框
            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk_item.setCheckState(Qt.Checked if item.get("selected", True) else Qt.Unchecked)
            chk_item.setData(Qt.UserRole, idx)
            self.table.setItem(row_pos, 0, chk_item)

            # 内容列
            cols = [
                d.get("nickname", ""),
                d.get("title", ""),
                d.get("bvid", ""),
                d.get("pub_time", ""),
                str(d.get("current_play", 0)),
                str(d.get("stat_days", 0)),
                d.get("play_tag", "")
            ]
            for c, text in enumerate(cols):
                ti = QTableWidgetItem(text)
                self.table.setItem(row_pos, c + 1, ti)
                if c + 1 != 2:
                    ti.setTextAlignment(Qt.AlignCenter)

            row_pos += 1

        # 4. 重新连接信号
        self.table.itemChanged.connect(self.on_item_check_changed)

    def on_item_check_changed(self, item):
        if item.column() != 0:
            return
        try:
            original_idx = item.data(Qt.UserRole)
            if original_idx is not None and 0 <= original_idx < len(self.data_list):
                self.data_list[original_idx]["selected"] = (item.checkState() == Qt.Checked)
        except Exception as e:
            pass

    def select_current_visible(self):
        current_up = self.up_combo.currentData()
        for item in self.data_list:
            d = item.get("data", {})
            if current_up is not None:
                target_uid, _ = current_up
                if d.get("uid", 0) != target_uid:
                    continue
            item["selected"] = True
        self.refresh_table_view()

    def deselect_current_visible(self):
        current_up = self.up_combo.currentData()
        for item in self.data_list:
            d = item.get("data", {})
            if current_up is not None:
                target_uid, _ = current_up
                if d.get("uid", 0) != target_uid:
                    continue
            item["selected"] = False
        self.refresh_table_view()

    def confirm_selection(self):
        # 1. 保存排除状态回历史文件
        daily = load_daily_video_data()
        for item in self.data_list:
            bvid = item.get("source_bvid", "") or item.get("data", {}).get("bvid", "")
            if bvid and bvid in daily:
                daily[bvid]["excluded"] = not item.get("selected", True)
        save_daily_video_data(daily)

        # 2. 准备返回给主窗口的数据
        self.result_data = [it["data"] for it in self.data_list if it.get("selected", True)]

        if not self.result_data:
            QMessageBox.warning(self, "提示", "至少保留一个视频！")
            return

        self.accept()
# -------------------------- Cookie设置对话框 --------------------------
class CookieSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Cookie 设置")
        self.setModal(True)
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)
        cc = load_cookie_config()
        self.sess = QLineEdit(cc.get("SESSDATA", ""))
        self.bili = QLineEdit(cc.get("BILI_JCT", ""))
        self.buvid = QLineEdit(cc.get("BUVID3", ""))
        self.dede = QLineEdit(cc.get("DEDEUSERID", ""))
        self.ac_time = QLineEdit(cc.get("AC_TIME_VALUE", "1000"))
        layout.addRow("DedeUserID:", self.dede)
        layout.addRow("ac_time_value:", self.ac_time)
        layout.addRow("SESSDATA:", self.sess)
        layout.addRow("bili_jct:", self.bili)
        layout.addRow("buvid3:", self.buvid)
        bl = QHBoxLayout()
        sav = QPushButton("保存")
        sav.clicked.connect(self.save_and_close)
        can = QPushButton("取消")
        can.clicked.connect(self.reject)
        bl.addStretch()
        bl.addWidget(sav)
        bl.addWidget(can)
        layout.addRow(bl)

    def save_and_close(self):
        save_cookie_config(
            self.sess.text().strip(),
            self.bili.text().strip(),
            self.buvid.text().strip(),
            self.dede.text().strip(),
            self.ac_time.text().strip()
        )
        # 更新全局变量（全部 5 个都要更新）
        global SESSDATA, BILI_JCT, BUVID3, DEDEUSERID, AC_TIME_VALUE
        SESSDATA = self.sess.text().strip()
        BILI_JCT = self.bili.text().strip()
        BUVID3 = self.buvid.text().strip()
        DEDEUSERID = self.dede.text().strip()
        AC_TIME_VALUE = self.ac_time.text().strip()

        QMessageBox.information(self, "成功", "Cookie已保存")
        self.accept()


# -------------------------- 主界面 --------------------------
class BiliStatTool(QMainWindow):
    def __init__(self):
        super().__init__()
        self.up_list = []
        self.stat_thread = None
        self.final_rank_data = []
        self.raw_video_data_backup = []
        self.init_ui()
        self.load_up_list()

    def init_ui(self):
        self.setWindowTitle("B站UP主7天播放量统计工具")
        self.setGeometry(100, 100, 1500, 850)
        self.setup_style()
        mw = QWidget()
        self.setCentralWidget(mw)
        ml = QVBoxLayout(mw)
        ml.setSpacing(15)
        ml.setContentsMargins(15, 15, 15, 15)

        cg = QGroupBox("统计配置")
        cl = QHBoxLayout(cg)
        cl.addWidget(QLabel("开始日期："))
        # self.sd = QDateEdit(QDate.currentDate().addMonths(-1))
        self.sd = QDateEdit(QDate.currentDate().addDays(-7))
        self.sd.setDisplayFormat("yyyy-MM-dd")
        self.sd.setCalendarPopup(True)
        cl.addWidget(self.sd)
        cl.addWidget(QLabel("结束日期："))
        self.ed = QDateEdit(QDate.currentDate())
        self.ed.setDisplayFormat("yyyy-MM-dd")
        self.ed.setCalendarPopup(True)
        cl.addWidget(self.ed)

        self.stb = QPushButton("开始统计")
        self.stb.setFixedSize(120, 35)
        self.stb.clicked.connect(self.start_stat)
        cl.addWidget(self.stb)

        self.spb = QPushButton("停止统计")
        self.spb.setFixedSize(120, 35)
        self.spb.clicked.connect(self.stop_stat)
        self.spb.setEnabled(False)
        cl.addWidget(self.spb)

        self.exb = QPushButton("导出Excel")
        self.exb.setFixedSize(120, 35)
        self.exb.clicked.connect(self.export_excel)
        self.exb.setEnabled(False)
        cl.addWidget(self.exb)

        self.fib = QPushButton("筛选作品")
        self.fib.setFixedSize(120, 35)
        self.fib.clicked.connect(self.open_filter_dialog)
        self.fib.setEnabled(False)
        cl.addWidget(self.fib)

        self.seb = QPushButton("Cookie设置")
        self.seb.setFixedSize(120, 35)
        self.seb.clicked.connect(self.open_settings)
        cl.addWidget(self.seb)

        self.heb = QPushButton("使用帮助")
        self.heb.setFixedSize(120, 35)
        self.heb.clicked.connect(self.open_help)
        cl.addWidget(self.heb)
        cl.addStretch()
        ml.addWidget(cg)

        sp = QSplitter(Qt.Horizontal)
        sp.setStretchFactor(0, 3)
        sp.setStretchFactor(1, 7)
        upw = QWidget()
        upl = QVBoxLayout(upw)
        upl.setContentsMargins(0, 0, 0, 0)
        upl.addWidget(QLabel(f"UP主管理（最多{MAX_UP_COUNT}个）"))
        aul = QHBoxLayout()
        aul.addWidget(QLabel("UID："))
        self.uid = QLineEdit()
        self.uid.setPlaceholderText("纯数字UID")
        aul.addWidget(self.uid)
        aul.addWidget(QLabel("昵称："))
        self.nk = QLineEdit()
        self.nk.setPlaceholderText("UP主昵称")
        aul.addWidget(self.nk)
        self.ab = QPushButton("添加")
        self.ab.clicked.connect(self.add_up)
        aul.addWidget(self.ab)
        upl.addLayout(aul)

        bal = QHBoxLayout()
        self.bt = QTextEdit()
        self.bt.setPlaceholderText("批量导入：\nUID,昵称")
        self.bt.setFixedHeight(100)
        bal.addWidget(self.bt)
        bbl = QVBoxLayout()
        self.ba = QPushButton("批量添加")
        self.ba.clicked.connect(self.batch_add_up)
        bbl.addWidget(self.ba)
        self.cb = QPushButton("清空列表")
        self.cb.clicked.connect(self.clear_up_list)
        bbl.addWidget(self.cb)
        bal.addLayout(bbl)
        upl.addLayout(bal)
        upl.addWidget(QLabel("已添加UP主："))
        self.ult = QTextEdit()
        self.ult.setReadOnly(True)
        upl.addWidget(self.ult)
        upw.setFixedWidth(420)
        sp.addWidget(upw)

        daw = QWidget()
        dal = QVBoxLayout(daw)
        dal.setContentsMargins(0, 0, 0, 0)
        dal.addWidget(QLabel("视频7天播放量统计详情："))
        # 【修改】主界面表格：新增统计天数字段
        self.vt = QTableWidget()
        self.vt.setColumnCount(7)
        self.vt.setHorizontalHeaderLabels(["UP主昵称", "视频标题", "BV号", "发布时间", "7天播放量", "统计天数", "状态"])
        self.vt.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.vt.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.vt.setEditTriggers(QTableWidget.NoEditTriggers)
        self.vt.setAlternatingRowColors(True)
        dal.addWidget(self.vt)
        dal.addWidget(QLabel("运行日志："))
        self.logt = QTextEdit()
        self.logt.setReadOnly(True)
        self.logt.setFixedHeight(150)
        dal.addWidget(self.logt)
        sp.addWidget(daw)
        ml.addWidget(sp)

    def setup_style(self):
        self.setStyleSheet("""
            QMainWindow{background-color:#f0f2f5;}
            QGroupBox{font-weight:bold;border:1px solid #d0d0d0;border-radius:6px;margin-top:12px;padding-top:12px;background:#fff;}
            QGroupBox::title{subcontrol-origin:margin;left:15px;padding:0 5px;color:#333;}
            QPushButton{background:#007AFF;color:white;border-radius:4px;padding:6px 12px;font-weight:bold;}
            QPushButton:hover{background:#58ACFA;}
            QPushButton:pressed{background:#045FB4;}
            QPushButton:disabled{background:#ccc;color:#666;}
            QLineEdit,QDateEdit,QTextEdit{border:1px solid #ccc;border-radius:4px;padding:5px;background:#fff;}
            QTableWidget{gridline-color:#e0e0e0;border:1px solid #d0d0d0;background:#fff;}
            QHeaderView::section{background:#f5f5f5;padding:6px;border:1px solid #d0d0d0;font-weight:bold;}
        """)

    def open_settings(self):
        CookieSettingsDialog(self).exec_()

    def open_help(self):
        QMessageBox.about(self, "使用帮助",
                          "1.添加UP主UID和昵称\n2.设置时间(默认是7天前)\n3.每日点击【开始统计】\n4.已结算视频自动锁定\n5.导出Excel\n\n⚠️ 重要提示：请务必在Cookie设置中填写完整的SESSDATA、bili_jct和buvid3！")

    def closeEvent(self, e):
        if self.stat_thread and self.stat_thread.isRunning():
            self.stat_thread.requestInterruption()
            self.stat_thread.wait(3000)
        e.accept()

    def load_up_list(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.up_list = json.load(f)
                self.up_list = [u for u in self.up_list if u.get("uid") and u.get("nickname")]
                self.refresh_up_list_text()
            except:
                self.up_list = []

    def save_up_list(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.up_list, f, ensure_ascii=False, indent=2)

    def refresh_up_list_text(self):
        t = ""
        for i, u in enumerate(self.up_list, 1):
            t += f"{i}.{u['nickname']}（UID:{u['uid']}）\n"
        self.ult.setText(t)

    def add_up(self):
        u = self.uid.text().strip()
        n = self.nk.text().strip()
        if not u or not n:
            QMessageBox.warning(self, "提示", "UID和昵称不能为空")
            return
        if not re.match(r'^\d+$', u):
            QMessageBox.warning(self, "提示", "UID必须纯数字")
            return
        if len(self.up_list) >= MAX_UP_COUNT:
            QMessageBox.warning(self, "提示", f"最多{MAX_UP_COUNT}个")
            return
        if any(str(x["uid"]) == str(u) for x in self.up_list):
            QMessageBox.warning(self, "提示", "已存在")
            return
        self.up_list.append({"uid": u, "nickname": n})
        self.save_up_list()
        self.refresh_up_list_text()
        self.uid.clear()
        self.nk.clear()
        self.log(f"已添加：{n}")

    def batch_add_up(self):
        t = self.bt.toPlainText().strip().replace("，", ",")
        if not t:
            QMessageBox.warning(self, "提示", "请输入内容")
            return
        c = 0
        for line in t.split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = line.split(",")
            if len(parts) != 2:
                continue
            uid, name = parts[0].strip(), parts[1].strip()
            if not re.match(r'^\d+$', uid) or not name:
                continue
            if len(self.up_list) >= MAX_UP_COUNT:
                break
            if any(str(x["uid"]) == str(uid) for x in self.up_list):
                continue
            self.up_list.append({"uid": uid, "nickname": name})
            c += 1
        if c > 0:
            self.save_up_list()
            self.refresh_up_list_text()
            self.bt.clear()
            self.log(f"批量添加{c}个")
        else:
            QMessageBox.warning(self, "提示", "无有效数据")

    def clear_up_list(self):
        if QMessageBox.question(self, "确认", "清空所有？") == QMessageBox.Yes:
            self.up_list = []
            self.save_up_list()
            self.refresh_up_list_text()
            self.log("已清空")

    def start_stat(self):
        if not self.up_list:
            QMessageBox.warning(self, "提示", "请添加UP主")
            return
        sdt = self.sd.date().toPyDate()
        edt = self.ed.date().toPyDate()
        if sdt > edt:
            QMessageBox.warning(self, "提示", "开始日期不能晚于结束日期")
            return
        sdt_full = datetime.combine(sdt, datetime.min.time(), tzinfo=TIMEZONE_CN)
        edt_full = datetime.combine(edt, datetime.max.time(), tzinfo=TIMEZONE_CN)
        now = datetime.now(tz=TIMEZONE_CN)
        if edt_full > now:
            edt_full = now
        self.vt.setRowCount(0)
        self.logt.clear()
        self.exb.setEnabled(False)
        self.fib.setEnabled(False)
        self.stb.setEnabled(False)
        self.spb.setEnabled(True)
        # 传入完整的Cookie参数

        self.stat_thread = StatThread(self.up_list, sdt_full, edt_full, SESSDATA, BILI_JCT, BUVID3, DEDEUSERID,
                                      AC_TIME_VALUE)
        self.stat_thread.log_signal.connect(self.log)
        self.stat_thread.table_signal.connect(self.refresh_table)
        self.stat_thread.finish_signal.connect(self.stat_finish)
        self.stat_thread.error_signal.connect(self.log)
        self.stat_thread.start()

    def stop_stat(self):
        if self.stat_thread and self.stat_thread.isRunning():
            self.stat_thread.requestInterruption()
            self.stat_thread.wait(2000)
            self.log("已停止")
        self.stb.setEnabled(True)
        self.spb.setEnabled(False)

    def refresh_table(self, data):
        self.vt.setRowCount(len(data))
        # 【修改】表格字段顺序
        keys = ["nickname", "title", "bvid", "pub_time", "current_play", "stat_days", "play_tag"]
        for row, item in enumerate(data):
            for col, key in enumerate(keys):
                val = str(item.get(key, ""))
                table_item = QTableWidgetItem(val)
                self.vt.setItem(row, col, table_item)
                if col != 1:
                    table_item.setTextAlignment(Qt.AlignCenter)
        self.vt.scrollToBottom()

    def stat_finish(self, rank_data):
        self.final_rank_data = rank_data
        self.raw_video_data_backup = []
        for row in range(self.vt.rowCount()):
            data = {}
            keys = ["nickname", "title", "bvid", "pub_time", "current_play", "stat_days", "play_tag"]
            for col, key in enumerate(keys):
                item = self.vt.item(row, col)
                data[key] = item.text() if item else ""
            try:
                for up in self.up_list:
                    if up["nickname"] == data["nickname"]:
                        data["uid"] = int(up["uid"])
                        break
                else:
                    data["uid"] = 0
            except:
                data["uid"] = 0
            try:
                data["current_play"] = int(data["current_play"])
                data["stat_days"] = int(data["stat_days"])
            except:
                data["current_play"] = 0
                data["stat_days"] = 0
            self.raw_video_data_backup.append(data)

        self.log("\n===== 统计完成 =====")
        self.log("✅ 已结算视频：数据永久锁定")
        self.log("🔄 统计中视频：次日将自动更新")
        for item in rank_data:
            self.log(f"{item['rank']}. {item['nickname']}：{item['total_play']}")
        self.stb.setEnabled(True)
        self.spb.setEnabled(False)
        self.exb.setEnabled(True)
        self.fib.setEnabled(True)
        QMessageBox.information(self, "完成", "统计完成！\n已结算视频锁定，统计中视频次日更新")
        # 新增：自动备份数据
        backup_dir = os.path.join(get_app_data_dir(), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        backup_file = os.path.join(backup_dir, f"video_7day_data_{datetime.now().strftime('%Y%m%d')}.json")
        try:
            shutil.copy(get_daily_data_path(), backup_file)
            self.log(f"✅ 数据已备份至：{backup_file}")
        except Exception as e:
            self.log(f"⚠️ 备份失败：{str(e)}")

    def export_excel(self):
        if self.vt.rowCount() == 0 or not self.final_rank_data:
            QMessageBox.warning(self, "提示", "无数据")
            return

        fp, _ = QFileDialog.getSaveFileName(
            self, "保存Excel",
            f"B站7天统计_完整历史数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel(*.xlsx)"
        )
        if not fp:
            return

        try:
            wb = openpyxl.Workbook()

            # ====================== 【1】视频明细（全部历史） ======================
            ws_detail = wb.active
            ws_detail.title = "视频明细（全部历史）"

            headers = ["UP主昵称", "视频标题", "BV号", "发布时间", "7天播放量", "统计天数", "统计状态"]
            for col, h in enumerate(headers, 1):
                cell = ws_detail.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            daily_data = load_daily_video_data()
            full_video_list = []
            seen_bvids = set()

            # 历史已结算视频
            for v in daily_data.values():
                if v.get("excluded", False):
                    continue
                if v.get("status") == "已结算" and "final_play" in v:
                    bvid = v.get("bvid")
                    if not bvid or bvid in seen_bvids:
                        continue
                    seen_bvids.add(bvid)
                    full_video_list.append({
                        "nickname": v.get("nickname", ""),
                        "title": v.get("title", ""),
                        "bvid": bvid,
                        "pub_time": v.get("pub_time", ""),
                        "current_play": v.get("final_play", 0),
                        "stat_days": v.get("stat_days", 7),
                        "play_tag": v.get("status", "已结算")
                    })

            # 本次统计中视频
            for video in self.raw_video_data_backup:
                if video.get("excluded", False):
                    continue
                if video.get("play_tag") == "统计中":
                    bvid = video.get("bvid")
                    if bvid and bvid not in seen_bvids:
                        seen_bvids.add(bvid)
                        full_video_list.append({
                            "nickname": video.get("nickname", ""),
                            "title": video.get("title", ""),
                            "bvid": bvid,
                            "pub_time": video.get("pub_time", ""),
                            "current_play": video.get("current_play", 0),
                            "stat_days": video.get("stat_days", 0),
                            "play_tag": video.get("play_tag", "统计中")
                        })

            # 按发布时间倒序
            full_video_list.sort(key=lambda x: x["pub_time"], reverse=True)

            # 填充数据（从第3行开始）
            up_names = [v["nickname"] for v in full_video_list]
            unique_ups = []
            seen = set()
            for name in up_names:
                if name and name not in seen:
                    seen.add(name)
                    unique_ups.append(name)

            colors = [PatternFill('solid', 'F2F2F2'), PatternFill('solid', 'E6F3FF'),
                      PatternFill('solid', 'FFF2E6'), PatternFill('solid', 'F0FFF0')]
            up_color = {name: colors[i % 4] for i, name in enumerate(unique_ups)}

            for row_idx, video in enumerate(full_video_list, 2):
                up_name = video["nickname"]
                fill = up_color.get(up_name, PatternFill('solid', 'FFFFFF'))

                values = [
                    video["nickname"], video["title"], video["bvid"],
                    video["pub_time"], str(video["current_play"]),
                    str(video["stat_days"]), video["play_tag"]
                ]

                for col, val in enumerate(values, 1):
                    cell = ws_detail.cell(row_idx, col, val)
                    cell.font = content_font
                    cell.border = thin_border
                    cell.alignment = left_align if col == 2 else center_align
                    if col == 5 and str(val).isdigit():
                        cell.number_format = "#,##0"
                    cell.fill = fill

            # 自动列宽 + 冻结窗格（冻结标题+表头）
            for col in range(1, 8):
                letter = get_column_letter(col)
                max_len = max((len(str(cell.value or "")) for cell in ws_detail[letter]), default=0)
                width = min(max_len + 8, 60) if col != 2 else min(max_len * 1.2 + 5, 80)
                ws_detail.column_dimensions[letter].width = width
            ws_detail.freeze_panes = "A2"

            # ====================== 【2】UP主整体排名（全部历史） ======================
            ws_rank = wb.create_sheet("UP主整体排名")

            rank_headers = ["档位", "排名", "UP主昵称", "UID", "7天总播放量（整体）"]
            for col, h in enumerate(rank_headers, 1):
                cell = ws_rank.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # 计算整体排名（修复：增加 excluded 过滤）
            settled_map = {}
            for v in daily_data.values():
                if v.get("excluded", False):  # 新增：过滤已排除视频
                    continue
                if v.get("status") == "已结算" and "final_play" in v:
                    uid = v.get("uid")
                    if uid:
                        if uid not in settled_map:
                            settled_map[uid] = {"nickname": v["nickname"], "uid": uid, "total_play": 0}
                        settled_map[uid]["total_play"] += v["final_play"]

            ongoing_map = {}
            for video in self.raw_video_data_backup:
                if video.get("excluded", False):  # 新增：过滤已排除视频
                    continue
                if video.get("play_tag") == "统计中":
                    uid = video.get("uid", 0)
                    if uid:
                        if uid not in ongoing_map:
                            ongoing_map[uid] = {"nickname": video["nickname"], "uid": uid, "total_play": 0}
                        ongoing_map[uid]["total_play"] += video.get("current_play", 0)

            merged_map = settled_map.copy()
            for uid, ongoing in ongoing_map.items():
                if uid in merged_map:
                    merged_map[uid]["total_play"] += ongoing["total_play"]
                else:
                    merged_map[uid] = ongoing.copy()

            merged_list = list(merged_map.values())
            merged_list.sort(key=lambda x: (-x["total_play"], x["nickname"]))
            for idx, item in enumerate(merged_list, 1):
                item["rank"] = idx

            # 档位数据（从第3行开始）
            current_row = 2
            normal_levels = RANK_LEVEL_CONFIG[:-1]
            last_level = RANK_LEVEL_CONFIG[-1]

            for level in normal_levels:
                name, start, end, color = level
                count = end - start + 1
                users = [u for u in merged_list if start <= u["rank"] <= end]
                merge_start = current_row
                merge_end = current_row + count - 1
                ws_rank.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
                level_fill = PatternFill('solid', fgColor=color)
                merged_cell = ws_rank.cell(merge_start, 1, name)
                merged_cell.font = Font('微软雅黑', 11, True, '000000')
                merged_cell.fill = level_fill
                merged_cell.alignment = center_align
                merged_cell.border = thin_border

                for i in range(count):
                    rank = start + i
                    user = users[i] if i < len(users) else None
                    ws_rank.cell(current_row, 2, rank)
                    if user:
                        ws_rank.cell(current_row, 3, user["nickname"])
                        ws_rank.cell(current_row, 4, str(user["uid"]))
                        ws_rank.cell(current_row, 5, user["total_play"]).number_format = "#,##0"
                    else:
                        ws_rank.cell(current_row, 3, "-")
                        ws_rank.cell(current_row, 4, "-")
                        ws_rank.cell(current_row, 5, "-")
                    for c in range(2, 6):
                        cell = ws_rank.cell(current_row, c)
                        cell.font = content_font
                        cell.border = thin_border
                        cell.alignment = center_align
                        cell.fill = level_fill
                    current_row += 1

            # TOP50+ 部分
            name, start, end, color = last_level
            users = [u for u in merged_list if u["rank"] >= start]
            if users:
                merge_start = current_row
                merge_end = current_row + len(users) - 1
                ws_rank.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
                level_fill = PatternFill('solid', fgColor=color)
                merged_cell = ws_rank.cell(merge_start, 1, name)
                merged_cell.font = Font('微软雅黑', 11, True, '000000')
                merged_cell.fill = level_fill
                merged_cell.alignment = center_align
                merged_cell.border = thin_border
                for user in users:
                    ws_rank.cell(current_row, 2, user["rank"])
                    ws_rank.cell(current_row, 3, user["nickname"])
                    ws_rank.cell(current_row, 4, str(user["uid"]))
                    ws_rank.cell(current_row, 5, user["total_play"]).number_format = "#,##0"
                    for c in range(2, 6):
                        cell = ws_rank.cell(current_row, c)
                        cell.font = content_font
                        cell.border = thin_border
                        cell.alignment = center_align
                        cell.fill = level_fill
                    current_row += 1

            # 自动列宽 + 冻结
            for col in range(1, 6):
                letter = get_column_letter(col)
                max_len = max((len(str(cell.value or "")) for cell in ws_rank[letter]), default=0)
                ws_rank.column_dimensions[letter].width = max_len + 8
            ws_rank.freeze_panes = "A2"

            # ====================== 保存 ======================
            wb.save(fp)
            QMessageBox.information(
                self, "导出成功",
                f"✅ 已导出完整历史统计数据！\n\n"
                f"· 视频明细（全部历史）\n"
                f"· UP主整体排名（全部历史）\n\n"
                f"文件已保存至：\n{fp}"
            )
            self.log(f"✅ 导出完整历史数据成功：{fp}")

        except Exception as e:
            QMessageBox.warning(self, "失败", f"导出失败：{str(e)}")
            self.log(f"导出失败：{e}")

    def open_filter_dialog(self):
        if not self.raw_video_data_backup:
            QMessageBox.warning(self, "提示", "暂无视频数据")
            return
        dialog = VideoSelectionDialog(self.raw_video_data_backup, self)
        if dialog.exec_() == QDialog.Accepted:
            # 读取筛选结果
            if hasattr(dialog, 'result_data'):
                filtered = dialog.result_data
            else:
                filtered = [it["data"] for it in dialog.data_list if it["selected"]]

            if not filtered:
                QMessageBox.warning(self, "提示", "筛选结果为空，无法更新")
                return

            # 【关键】更新主窗口的备份数据，确保导出用的是新的
            self.raw_video_data_backup = filtered

            # 刷新当前界面表格
            self.refresh_table(filtered)

            # 重新计算排名
            total_map = {}
            for video in filtered:
                uid = video.get("uid", 0)
                nick = video.get("nickname", "")
                try:
                    play = int(video.get("current_play", 0))
                except:
                    play = 0
                if uid not in total_map:
                    total_map[uid] = {"uid": uid, "nickname": nick, "total_play": 0}
                total_map[uid]["total_play"] += play

            ranked = sorted(total_map.values(), key=lambda x: (-x["total_play"], x["nickname"]))
            for i, item in enumerate(ranked, 1):
                item["rank"] = i
            self.final_rank_data = ranked

            self.log(f"\n===== 筛选完成 =====")
            self.log(f"已保留 {len(filtered)} 个视频")
            for item in ranked[:10]:  # 只打印前10名
                self.log(f"{item['rank']}. {item['nickname']}：{item['total_play']}")
            QMessageBox.information(self, "完成", f"筛选完成！\n当前显示 {len(filtered)} 个视频")

    def log(self, text):
        self.logt.append(f"[{datetime.now().strftime('%H:%M:%S')}] {text}")
        self.logt.verticalScrollBar().setValue(self.logt.verticalScrollBar().maximum())


# -------------------------- 全局异常 --------------------------
def except_hook(t, v, tb):
    if issubclass(t, KeyboardInterrupt):
        sys.__excepthook__(t, v, tb)
        return
    log_dir = os.path.join(get_app_data_dir(), "logs")
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"crash_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write("".join(traceback.format_exception(t, v, tb)))
    try:
        QMessageBox.critical(None, "程序异常", f"错误已保存至：\n{log_file}")
    except:
        pass


sys.excepthook = except_hook

# -------------------------- 启动 --------------------------
if __name__ == "__main__":
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    app = QApplication(sys.argv)
    app.setFont(QFont("Microsoft YaHei", 9))
    window = BiliStatTool()
    window.show()
    sys.exit(app.exec_())

import os
from datetime import datetime
from PyQt5.QtWidgets import QFileDialog, QMessageBox
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import (
    RANK_LEVEL_CONFIG, thin_border, center_align, left_align,
    header_font, content_font, header_fill
)
from .storage import load_daily_video_data


def _sanitize_excel_value(val):
    """防止 Excel 公式注入：对以 = + - @ 开头的字符串添加 ' 前缀"""
    if isinstance(val, str) and val and val[0] in ('=', '+', '-', '@'):
        return "'" + val
    return val


def _cjk_len(s):
    """估算字符串显示宽度，CJK 字符按2倍计算"""
    s = str(s or "")
    return sum(2 if '一' <= c <= '鿿' or '　' <= c <= '〿' else 1 for c in s)


def export_excel(parent, raw_video_data_backup, final_rank_data, enabled_uids=None, settlement_data=None):
    if not raw_video_data_backup or not final_rank_data:
        QMessageBox.warning(parent, "提示", "无数据")
        return

    fp, _ = QFileDialog.getSaveFileName(
        parent, "保存Excel",
        f"B站7天统计_完整历史数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "Excel(*.xlsx)"
    )
    if not fp:
        return

    try:
        wb = openpyxl.Workbook()

        # ====================== 【1】视频明细（全部历史） ======================
        ws_detail = wb.active
        ws_detail.title = "视频明细（全部历史）"

        headers = ["UP主昵称", "视频标题", "BV号", "发布时间", "7天播放量", "统计天数", "统计状态", "共创角色"]
        for col, h in enumerate(headers, 1):
            cell = ws_detail.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        daily_data = load_daily_video_data()
        full_video_list = []
        seen_bvids = set()

        for v in daily_data.values():
            if v.get("excluded", False):
                continue
            if enabled_uids is not None and v.get("uid", 0) not in enabled_uids:
                continue
            if v.get("status") == "已结算" and "final_play" in v:
                bvid = v.get("bvid")
                if not bvid or bvid in seen_bvids:
                    continue
                seen_bvids.add(bvid)
                full_video_list.append({
                    "nickname": v.get("nickname", ""),
                    "title": v.get("title", ""),
                    "bvid": bvid,
                    "pub_time": v.get("pub_time", ""),
                    "current_play": v.get("final_play", 0),
                    "stat_days": v.get("stat_days", 7),
                    "play_tag": v.get("status", "已结算"),
                    "collab_role": v.get("collab_role", ""),
                })

        for video in raw_video_data_backup:
            if video.get("excluded", False):
                continue
            if video.get("play_tag") == "统计中":
                bvid = video.get("bvid")
                if bvid and bvid not in seen_bvids:
                    seen_bvids.add(bvid)
                    full_video_list.append({
                        "nickname": video.get("nickname", ""),
                        "title": video.get("title", ""),
                        "bvid": bvid,
                        "pub_time": video.get("pub_time", ""),
                        "current_play": video.get("current_play", 0),
                        "stat_days": video.get("stat_days", 0),
                        "play_tag": video.get("play_tag", "统计中"),
                        "collab_role": video.get("collab_role", ""),
                    })

        full_video_list.sort(key=lambda x: x["pub_time"], reverse=True)

        up_names = [v["nickname"] for v in full_video_list]
        unique_ups = []
        seen = set()
        for name in up_names:
            if name and name not in seen:
                seen.add(name)
                unique_ups.append(name)

        colors = [PatternFill('solid', 'F2F2F2'), PatternFill('solid', 'E6F3FF'),
                  PatternFill('solid', 'FFF2E6'), PatternFill('solid', 'F0FFF0')]
        up_color = {name: colors[i % 4] for i, name in enumerate(unique_ups)}

        for row_idx, video in enumerate(full_video_list, 2):
            up_name = video["nickname"]
            fill = up_color.get(up_name, PatternFill('solid', 'FFFFFF'))

            values = [
                video["nickname"], video["title"], video["bvid"],
                video["pub_time"], str(video["current_play"]),
                str(video["stat_days"]), video["play_tag"],
                video.get("collab_role", ""),
            ]

            for col, val in enumerate(values, 1):
                cell = ws_detail.cell(row_idx, col, val)
                cell.font = content_font
                cell.border = thin_border
                cell.alignment = left_align if col == 2 else center_align
                if col == 5 and str(val).isdigit():
                    cell.number_format = "#,##0"
                cell.fill = fill

        for col in range(1, 9):
            letter = get_column_letter(col)
            max_len = max((len(str(cell.value or "")) for cell in ws_detail[letter]), default=0)
            width = min(max_len + 8, 60) if col != 2 else min(max_len * 1.2 + 5, 80)
            ws_detail.column_dimensions[letter].width = width
        ws_detail.freeze_panes = "A2"

        # ====================== 【2】UP主整体排名（全部历史） ======================
        ws_rank = wb.create_sheet("UP主整体排名")

        rank_headers = ["档位", "排名", "UP主昵称", "UID", "7天总播放量（整体）", "视频数量"]
        for col, h in enumerate(rank_headers, 1):
            cell = ws_rank.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        settled_map = {}
        for v in daily_data.values():
            if v.get("excluded", False):
                continue
            if enabled_uids is not None and v.get("uid", 0) not in enabled_uids:
                continue
            if v.get("status") == "已结算" and "final_play" in v:
                uid = v.get("uid")
                if uid:
                    if uid not in settled_map:
                        settled_map[uid] = {"nickname": v["nickname"], "uid": uid, "total_play": 0, "video_count": 0}
                    settled_map[uid]["total_play"] += v["final_play"]
                    settled_map[uid]["video_count"] += 1

        ongoing_map = {}
        for video in raw_video_data_backup:
            if video.get("excluded", False):
                continue
            if video.get("play_tag") == "统计中":
                uid = video.get("uid", 0)
                if uid:
                    if uid not in ongoing_map:
                        ongoing_map[uid] = {"nickname": video["nickname"], "uid": uid, "total_play": 0, "video_count": 0}
                    ongoing_map[uid]["total_play"] += video.get("current_play", 0)
                    ongoing_map[uid]["video_count"] += 1

        merged_map = settled_map.copy()
        for uid, ongoing in ongoing_map.items():
            if uid in merged_map:
                merged_map[uid]["total_play"] += ongoing["total_play"]
                merged_map[uid]["video_count"] += ongoing["video_count"]
            else:
                merged_map[uid] = ongoing.copy()

        merged_list = list(merged_map.values())
        merged_list.sort(key=lambda x: (-x["total_play"], x["nickname"]))
        for idx, item in enumerate(merged_list, 1):
            item["rank"] = idx

        current_row = 2
        normal_levels = RANK_LEVEL_CONFIG[:-1]
        last_level = RANK_LEVEL_CONFIG[-1]

        for level in normal_levels:
            name, start, end, color = level
            count = end - start + 1
            users = [u for u in merged_list if start <= u["rank"] <= end]
            merge_start = current_row
            merge_end = current_row + count - 1
            ws_rank.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
            level_fill = PatternFill('solid', fgColor=color)
            merged_cell = ws_rank.cell(merge_start, 1, name)
            merged_cell.font = Font('微软雅黑', 11, True, '000000')
            merged_cell.fill = level_fill
            merged_cell.alignment = center_align
            merged_cell.border = thin_border

            for i in range(count):
                rank = start + i
                user = users[i] if i < len(users) else None
                ws_rank.cell(current_row, 2, rank)
                if user:
                    ws_rank.cell(current_row, 3, user["nickname"])
                    ws_rank.cell(current_row, 4, str(user["uid"]))
                    ws_rank.cell(current_row, 5, user["total_play"]).number_format = "#,##0"
                    ws_rank.cell(current_row, 6, user["video_count"])
                else:
                    ws_rank.cell(current_row, 3, "-")
                    ws_rank.cell(current_row, 4, "-")
                    ws_rank.cell(current_row, 5, "-")
                    ws_rank.cell(current_row, 6, "-")
                for c in range(2, 7):
                    cell = ws_rank.cell(current_row, c)
                    cell.font = content_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    cell.fill = level_fill
                current_row += 1

        name, start, end, color = last_level
        users = [u for u in merged_list if u["rank"] >= start]
        if users:
            merge_start = current_row
            merge_end = current_row + len(users) - 1
            ws_rank.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
            level_fill = PatternFill('solid', fgColor=color)
            merged_cell = ws_rank.cell(merge_start, 1, name)
            merged_cell.font = Font('微软雅黑', 11, True, '000000')
            merged_cell.fill = level_fill
            merged_cell.alignment = center_align
            merged_cell.border = thin_border
            for user in users:
                ws_rank.cell(current_row, 2, user["rank"])
                ws_rank.cell(current_row, 3, user["nickname"])
                ws_rank.cell(current_row, 4, str(user["uid"]))
                ws_rank.cell(current_row, 5, user["total_play"]).number_format = "#,##0"
                ws_rank.cell(current_row, 6, user["video_count"])
                for c in range(2, 7):
                    cell = ws_rank.cell(current_row, c)
                    cell.font = content_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    cell.fill = level_fill
                current_row += 1

        for col in range(1, 7):
            letter = get_column_letter(col)
            max_len = max((len(str(cell.value or "")) for cell in ws_rank[letter]), default=0)
            ws_rank.column_dimensions[letter].width = max_len + 8
        ws_rank.freeze_panes = "A2"

        # ====================== 【3】奖池结算 ======================
        if settlement_data:
            ws_st = wb.create_sheet("奖池结算")

            BASE_COL = 7  # 从 G 列开始居中

            tier_colors = [color for _, _, _, color in RANK_LEVEL_CONFIG]
            tier_name_font = Font(name='微软雅黑', size=12, bold=True)
            sub_header_font = Font(name='微软雅黑', size=10, bold=True, color='333333')
            row = 1
            tier_results = settlement_data.get("tier_results", [])

            for idx, tr in enumerate(tier_results):
                color = tier_colors[idx % len(tier_colors)]
                fill = PatternFill('solid', fgColor=color)

                # ── 档位名称行 ──
                ws_st.merge_cells(start_row=row, start_column=BASE_COL, end_row=row, end_column=BASE_COL + 2)
                tier_cell = ws_st.cell(row, BASE_COL, f"[{tr['tier_name']}]")
                tier_cell.font = tier_name_font
                tier_cell.fill = fill
                tier_cell.alignment = center_align
                tier_cell.border = thin_border
                for c in (BASE_COL + 1, BASE_COL + 2):
                    cell = ws_st.cell(row, c)
                    cell.fill = fill
                    cell.border = thin_border
                row += 1

                # ── 子表头行1：累计播放量 | 瓜分奖池 | 单人最高瓜分 ──
                sub_headers_1 = ["累计播放量", "瓜分奖池", "单人最高瓜分"]
                for i, h in enumerate(sub_headers_1):
                    cell = ws_st.cell(row, BASE_COL + i, h)
                    cell.font = sub_header_font
                    cell.fill = fill
                    cell.alignment = center_align
                    cell.border = thin_border
                row += 1

                # ── 数值行 ──
                vals_1 = [tr["threshold"], tr["pool"], tr["max_per_person"]]
                for i, val in enumerate(vals_1):
                    cell = ws_st.cell(row, BASE_COL + i, val)
                    cell.font = content_font
                    cell.fill = fill
                    cell.alignment = center_align
                    cell.border = thin_border
                    if i == 0:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.00"
                row += 1

                # ── 子表头行2：已达到的作者 | 每位作者的播放量 | 作者瓜分金额 ──
                sub_headers_2 = ["已达到的作者", "每位作者的播放量", "作者瓜分金额"]
                for i, h in enumerate(sub_headers_2):
                    cell = ws_st.cell(row, BASE_COL + i, h)
                    cell.font = sub_header_font
                    cell.fill = fill
                    cell.alignment = center_align
                    cell.border = thin_border
                row += 1

                # ── 作者明细行 ──
                for m in tr.get("members", []):
                    detail_vals = [m["nickname"], m["individual_views"], m["actual"]]
                    for i, val in enumerate(detail_vals):
                        cell = ws_st.cell(row, BASE_COL + i, val)
                        cell.font = content_font
                        cell.fill = fill
                        cell.alignment = center_align
                        cell.border = thin_border
                        if i == 1:
                            cell.number_format = "#,##0"
                        elif i == 2:
                            cell.number_format = "#,##0.00"
                    row += 1

                # ── 档位间空行 ──
                row += 1

            col_widths = {BASE_COL: 22, BASE_COL + 1: 22, BASE_COL + 2: 20}
            for col, w in col_widths.items():
                ws_st.column_dimensions[get_column_letter(col)].width = w
            ws_st.freeze_panes = "A2"

        # ====================== 保存 ======================
        wb.save(fp)
        sheet_names = "· 视频明细（全部历史）\n· UP主整体排名（全部历史）"
        if settlement_data:
            sheet_names += "\n· 奖池结算"
        QMessageBox.information(
            parent, "导出成功",
            f"✅ 已导出完整历史统计数据！\n\n"
            f"{sheet_names}\n\n"
            f"文件已保存至：\n{fp}"
        )
        return fp

    except Exception as e:
        QMessageBox.warning(parent, "失败", f"导出失败：{str(e)}")
        return None

import os
import json
from datetime import timedelta, timezone
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# -------------------------- 工具函数：获取用户数据目录 --------------------------
def get_app_data_dir():
    home_dir = os.path.expanduser("~")
    app_dir = os.path.join(home_dir, ".BiliStatTool")
    if not os.path.exists(app_dir):
        os.makedirs(app_dir)
    return app_dir


# -------------------------- Cookie 配置 --------------------------
def load_cookie_config():
    cookie_file = os.path.join(get_app_data_dir(), "cookie.json")
    if os.path.exists(cookie_file):
        try:
            with open(cookie_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "SESSDATA": "",
        "BILI_JCT": "",
        "BUVID3": "",
        "DEDEUSERID": "",
        "AC_TIME_VALUE": "1000"
    }


def save_cookie_config(sessdata, bili_jct, buvid3, dedeuserid, ac_time_value):
    cookie_file = os.path.join(get_app_data_dir(), "cookie.json")
    with open(cookie_file, "w", encoding="utf-8") as f:
        json.dump({
            "SESSDATA": sessdata,
            "BILI_JCT": bili_jct,
            "BUVID3": buvid3,
            "DEDEUSERID": dedeuserid,
            "AC_TIME_VALUE": ac_time_value
        }, f, ensure_ascii=False, indent=2)


def update_cookie(sessdata, bili_jct, buvid3, dedeuserid, ac_time_value):
    save_cookie_config(sessdata, bili_jct, buvid3, dedeuserid, ac_time_value)
    global SESSDATA, BILI_JCT, BUVID3, DEDEUSERID, AC_TIME_VALUE
    SESSDATA = sessdata
    BILI_JCT = bili_jct
    BUVID3 = buvid3
    DEDEUSERID = dedeuserid
    AC_TIME_VALUE = ac_time_value


# -------------------------- 全局配置 --------------------------
cookie_config = load_cookie_config()
SESSDATA = cookie_config.get("SESSDATA", "")
BILI_JCT = cookie_config.get("BILI_JCT", "")
BUVID3 = cookie_config.get("BUVID3", "")
DEDEUSERID = cookie_config.get("DEDEUSERID", "")
AC_TIME_VALUE = cookie_config.get("AC_TIME_VALUE", "1000")

MAX_UP_COUNT = 80
CONFIG_FILE = os.path.join(get_app_data_dir(), "up_list.json")
SETTINGS_FILE = os.path.join(get_app_data_dir(), "settings.json")
TIMEZONE_CN = timezone(timedelta(hours=8), "Asia/Shanghai")


def load_role_filter_settings():
    """加载共创角色筛选设置，返回 (enabled: bool, roles: list[str])，默认 (False, [])"""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                s = json.load(f)
                return s.get("role_filter_enabled", False), s.get("excluded_roles", [])
        except Exception:
            pass
    return False, []


def save_role_filter_settings(enabled, roles):
    """保存共创角色筛选设置到 settings.json"""
    settings = {}
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                settings = json.load(f)
        except Exception:
            pass
    settings["role_filter_enabled"] = enabled
    settings["excluded_roles"] = roles
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

RANK_LEVEL_CONFIG = [
    ("TOP1", 1, 1, "D6C7FB"),
    ("TOP2-3", 2, 3, "FEDBF5"),
    ("TOP4-5", 4, 5, "FEF7FF"),
    ("TOP6-10", 6, 10, "E0EEFF"),
    ("TOP11-15", 11, 15, "FFFDE5"),
    ("TOP16-20", 16, 20, "F5F891"),
    ("TOP21-25", 21, 25, "D1F9BA"),
    ("TOP26-30", 26, 30, "72A1A7"),
    ("TOP31-35", 31, 35, "BA9F7D"),
    ("TOP36-40", 36, 40, "C4D69F"),
    ("TOP41-50", 41, 50, "FFBE4E"),
    ("TOP50+", 51, 999, "9e9e9e"),
]

# -------------------------- Excel 样式常量 --------------------------
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
content_font = Font(name='微软雅黑', size=10)
header_fill = PatternFill('solid', fgColor='4472C4')
odd_row_fill = PatternFill('solid', fgColor='F2F2F2')
even_row_fill = PatternFill('solid', fgColor='FFFFFF')
